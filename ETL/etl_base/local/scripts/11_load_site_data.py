"""
SKALA Physical Risk AI System - 사이트 추가 데이터 적재 (범용 버전)
어떤 Excel 파일이든 텍스트 덤프로 저장하여 LLM이 나중에 활용

데이터 소스:
    - data 폴더 내 모든 *.xlsx 파일
대상 테이블:
    - site_additional_data (범용 JSONB 저장)

최종 수정일: 2025-12-10
버전: v03 (범용 버전)
"""

import sys
import os
import json
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook

from utils import setup_logging, get_db_connection, get_data_dir, table_exists, get_row_count


# 고정 Site ID (환경변수 또는 기본값)
PANGYO_DC_SITE_ID = os.environ.get('PANGYO_DC_SITE_ID', '00000000-0000-0000-0000-000000000001')
PANGYO_CAMPUS_SITE_ID = os.environ.get('PANGYO_CAMPUS_SITE_ID', '00000000-0000-0000-0000-000000000002')
DEFAULT_SITE_ID = os.environ.get('DEFAULT_SITE_ID', '00000000-0000-0000-0000-000000000099')


def universal_excel_loader(xlsx_file: Path) -> dict:
    """
    어떤 Excel이든 텍스트 덤프로 변환
    - 병합 셀, 빈 행, 이상한 구조 모두 OK
    - LLM이 나중에 해석

    Args:
        xlsx_file: Excel 파일 경로

    Returns:
        dict: 텍스트 덤프된 데이터
    """
    result = {
        'file_name': xlsx_file.name,
        'uploaded_at': datetime.now().isoformat(),
        'sheets': []
    }

    try:
        wb = load_workbook(xlsx_file, data_only=True)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            rows = []
            for row in ws.iter_rows():
                row_values = [str(cell.value) if cell.value is not None else '' for cell in row]
                # 완전히 빈 행은 스킵
                if any(v.strip() for v in row_values):
                    rows.append(' | '.join(row_values))

            result['sheets'].append({
                'name': sheet_name,
                'row_count': len(rows),
                'content': '\n'.join(rows)
            })

        wb.close()

    except Exception as e:
        result['error'] = str(e)
        result['sheets'] = []

    return result


def infer_category(file_name: str) -> str:
    """
    파일명에서 카테고리 추론

    Args:
        file_name: 파일명

    Returns:
        str: 추론된 카테고리
    """
    file_lower = file_name.lower()

    if '전력' in file_lower or 'power' in file_lower:
        return 'power'
    elif '에너지' in file_lower or 'energy' in file_lower:
        return 'energy'
    elif '보험' in file_lower or 'insurance' in file_lower:
        return 'insurance'
    elif '건물' in file_lower or 'building' in file_lower:
        return 'building'
    elif '자산' in file_lower or 'asset' in file_lower:
        return 'asset'
    else:
        return 'other'


def infer_site_id(file_name: str) -> str:
    """
    파일명에서 site_id 추론

    Args:
        file_name: 파일명

    Returns:
        str: 추론된 site_id
    """
    file_lower = file_name.lower()

    if '판교dc' in file_lower or '판교DC' in file_name:
        return PANGYO_DC_SITE_ID
    elif '판교캠퍼스' in file_lower:
        return PANGYO_CAMPUS_SITE_ID
    else:
        return DEFAULT_SITE_ID


def load_site_data() -> None:
    """data 폴더의 모든 Excel 파일을 범용 방식으로 로드"""
    logger = setup_logging("load_site_data")
    logger.info("=" * 60)
    logger.info("사이트 추가 데이터 로딩 시작 (범용 버전)")
    logger.info("=" * 60)

    # DB 연결 테스트
    try:
        conn = get_db_connection()
        logger.info("데이터베이스 연결 성공")
    except Exception as e:
        logger.error(f"데이터베이스 연결 실패: {e}")
        sys.exit(1)

    # 테이블 존재 확인
    if not table_exists(conn, "site_additional_data"):
        logger.error("site_additional_data 테이블이 존재하지 않습니다")
        conn.close()
        sys.exit(1)

    cursor = conn.cursor()
    data_dir = get_data_dir()

    # 모든 Excel 파일 찾기 (참조 데이터 파일 제외)
    EXCLUDE_PATTERNS = ['aqueduct', 'rankings', 'reference']  # 글로벌 참조 데이터 제외

    all_xlsx = list(data_dir.glob("*.xlsx"))
    xlsx_files = [f for f in all_xlsx if not any(p in f.name.lower() for p in EXCLUDE_PATTERNS)]

    logger.info(f"발견된 Excel 파일: {len(all_xlsx)}개 (제외: {len(all_xlsx) - len(xlsx_files)}개)")

    if not xlsx_files:
        logger.warning("Excel 파일이 없습니다")
        conn.close()
        return

    success_count = 0
    fail_count = 0

    for xlsx_file in xlsx_files:
        logger.info(f"처리 중: {xlsx_file.name}")

        # 파일명으로 카테고리/사이트 추론
        category = infer_category(xlsx_file.name)
        site_id = infer_site_id(xlsx_file.name)

        logger.info(f"   → site_id: {site_id[:8]}..., category: {category}")

        # 범용 로더로 텍스트 덤프
        data = universal_excel_loader(xlsx_file)

        if 'error' in data:
            logger.error(f"   ✗ 파일 읽기 실패: {data['error']}")
            fail_count += 1
            continue

        # DB 저장 (UPSERT - site_id + category + file_name 기준)
        try:
            cursor.execute("""
                INSERT INTO site_additional_data
                (site_id, data_category, file_name, structured_data, metadata, uploaded_at)
                VALUES (%s, %s, %s, %s, %s, NOW())
                ON CONFLICT (site_id, data_category, file_name)
                DO UPDATE SET
                    structured_data = EXCLUDED.structured_data,
                    metadata = EXCLUDED.metadata,
                    uploaded_at = NOW()
            """, (
                site_id,
                category,
                xlsx_file.name,
                json.dumps(data, ensure_ascii=False),
                json.dumps({
                    'source': 'Universal Excel Loader',
                    'loaded_at': datetime.now().isoformat(),
                    'sheet_count': len(data['sheets']),
                    'total_rows': sum(s['row_count'] for s in data['sheets'])
                }, ensure_ascii=False)
            ))
            conn.commit()

            total_rows = sum(s['row_count'] for s in data['sheets'])
            logger.info(f"   ✓ 저장 완료: {len(data['sheets'])} sheets, {total_rows} rows")
            success_count += 1

        except Exception as e:
            logger.error(f"   ✗ DB 저장 실패: {e}")
            conn.rollback()
            fail_count += 1

    cursor.close()
    conn.close()

    # 결과 출력
    logger.info("=" * 60)
    logger.info("사이트 추가 데이터 로딩 완료")
    logger.info(f"   - 성공: {success_count}개 파일")
    logger.info(f"   - 실패: {fail_count}개 파일")
    logger.info("=" * 60)


if __name__ == "__main__":
    load_site_data()
